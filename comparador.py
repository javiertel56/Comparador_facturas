import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
import os # Importar os para abrir archivos

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Comparador de Facturas")
        self.geometry("700x500")
        self.configure(bg="#1c2c47") # Fondo oscuro para la ventana principal

        self.file1 = None
        self.file2 = None
        self.archivo_procesado = None
        
        # --- Configuración de Estilos ttk ---
        self.style = ttk.Style(self) # Inicializa el estilo ttk
        
        # Configurar el tema base de ttk (opcional, pero 'clam' o 'alt' suelen ser mejores para personalizar)
        self.style.theme_use('clam') 

        # Estilo para el título
        self.style.configure('TLabel', background="#1c2c47", foreground="#f5f6fa", font=('Segoe UI', 10))
        self.style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'))

        # Estilos para los botones
        # Botones "Subir archivo" (azul/gris-azulado)
        self.style.configure('Upload.TButton',
                             background='#4A90E2',  # Un azul más vibrante
                             foreground='white',
                             font=('Segoe UI', 10, 'bold'),
                             padding=[15, 10], # Mayor padding para botones más grandes
                             borderwidth=0,     # Sin borde visible por defecto
                             relief="flat"      # Estilo plano
                            )
        self.style.map('Upload.TButton',
                       background=[('active', '#357ABD'), # Azul oscuro al pasar el ratón
                                   ('pressed', '#286090')], # Azul aún más oscuro al hacer clic
                       foreground=[('active', 'white'), 
                                   ('pressed', 'white')]
                      )

        # Botón "Comparar" (Amarillo/Naranja)
        self.style.configure('Compare.TButton',
                             background='#F5A623', # Naranja brillante
                             foreground='white',
                             font=('Segoe UI', 10, 'bold'),
                             padding=[15, 10],
                             borderwidth=0,
                             relief="flat"
                            )
        self.style.map('Compare.TButton',
                       background=[('active', '#E09117'), 
                                   ('pressed', '#C97D00')],
                       foreground=[('active', 'white'), 
                                   ('pressed', 'white')]
                      )

        # Botón "Abrir resultado" (Verde)
        self.style.configure('Open.TButton',
                             background='#50E3C2', # Verde aguamarina
                             foreground='#1c2c47', # Texto oscuro para contraste
                             font=('Segoe UI', 10, 'bold'),
                             padding=[15, 10],
                             borderwidth=0,
                             relief="flat"
                            )
        self.style.map('Open.TButton',
                       background=[('active', '#36B69D'), 
                                   ('pressed', '#2A8E77')],
                       foreground=[('active', '#1c2c47'), 
                                   ('pressed', '#1c2c47')]
                      )

        self.create_content()

    def create_content(self):
        # Usamos el estilo 'Title.TLabel' aquí
        lbl_titulo = ttk.Label(self, text="Comparador de Facturas SAT vs Odoo", style='Title.TLabel')
        lbl_titulo.pack(pady=(20, 20))

        # Es mejor usar ttk.Frame para que se apliquen los estilos de ttk si se usan temas
        btn_frame = ttk.Frame(self, style='TFrame') # Puedes definir un estilo para TFrame si quieres bg
        btn_frame.pack(pady=10)

        # Configuramos el background del frame de botones manualmente si ttk.Frame no hereda bien el bg
        # O si prefieres un color diferente para el frame de botones
        btn_frame.configure(style='Custom.TFrame') 
        self.style.configure('Custom.TFrame', background="#1c2c47")

        self.lbl_file1 = ttk.Label(self, text="Archivo 1 (Odoo): Ninguno", style='TLabel')
        self.lbl_file1.pack(pady=(5, 5))
        self.lbl_file2 = ttk.Label(self, text="Archivo 2 (SAT): Ninguno", style='TLabel')
        self.lbl_file2.pack(pady=(5, 20))

        # --- Funciones de Botones (sin cambios en la lógica) ---
        def subir_archivo1():
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
            if file_path:
                self.file1 = file_path
                self.lbl_file1.config(text=f"Archivo 1 (Odoo): {os.path.basename(file_path)}") # Mostrar solo el nombre del archivo
        
        def subir_archivo2():
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
            if file_path:
                self.file2 = file_path
                self.lbl_file2.config(text=f"Archivo 2 (SAT): {os.path.basename(file_path)}") # Mostrar solo el nombre del archivo

        def procesar_excel():
            if not self.file1 or not self.file2:
                messagebox.showwarning("Advertencia", "Debes seleccionar ambos archivos.")
                return
            try:
                destino = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                if not destino:
                    return
                comparar_facturas(self.file1, self.file2, destino)
                self.archivo_procesado = destino
                messagebox.showinfo("Éxito", "Comparación realizada y guardada correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error: {e}")

        def abrir_archivo():
            if hasattr(self, 'archivo_procesado') and self.archivo_procesado and os.path.exists(self.archivo_procesado):
                try:
                    os.startfile(self.archivo_procesado)
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
            else:
                messagebox.showwarning("Advertencia", "Primero procesa y guarda un archivo.")

        # --- Creación de Botones con los nuevos estilos ---
        btn_subir1 = ttk.Button(btn_frame, text="Subir archivo 1 (Odoo)", style="Upload.TButton", command=subir_archivo1)
        btn_subir1.pack(side=tk.LEFT, padx=10)
        
        btn_subir2 = ttk.Button(btn_frame, text="Subir archivo 2 (SAT)", style="Upload.TButton", command=subir_archivo2)
        btn_subir2.pack(side=tk.LEFT, padx=10)
        
        btn_procesar = ttk.Button(btn_frame, text="Comparar", style="Compare.TButton", command=procesar_excel)
        btn_procesar.pack(side=tk.LEFT, padx=10)
        
        btn_abrir = ttk.Button(btn_frame, text="Abrir resultado", style="Open.TButton", command=abrir_archivo)
        btn_abrir.pack(side=tk.LEFT, padx=10)

# --- Las funciones de procesamiento de datos (sin cambios importantes, solo la importación de os) ---

def convertir_excel(archivo_entrada, archivo_salida, tipo):
    # Aquí va la lógica real de conversión
    # Por ahora solo copia el archivo de entrada al de salida
    import shutil
    shutil.copyfile(archivo_entrada, archivo_salida)
    ajustar_formato_excel(archivo_salida)

def limpiar_valor(valor):
    if pd.isna(valor):
        return 0.0
    try:
        return float(str(valor).replace(',', '').replace(' ', ''))
    except Exception:
        return 0.0

def ajustar_formato_excel(ruta_archivo):
    wb = openpyxl.load_workbook(ruta_archivo)
    azul = PatternFill(start_color="2980b9", end_color="2980b9", fill_type="solid")
    for ws in wb.worksheets:
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Encabezado azul, negrita y centrado
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = azul

        # Congelar la primera fila
        ws.freeze_panes = "A2"

        # Formato numérico y pintar de blanco los ceros en columnas de dinero e IVA
        header = [cell.value for cell in ws[1]]
        cols_dinero = []
        for nombre in ['Cargo 16', 'Abono 16', 'Cargo 8', 'Abono 8']:
            if nombre in header:
                cols_dinero.append(header.index(nombre) + 1)
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for idx in cols_dinero:
                cell = row[idx-1]
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                if cell.value == 0 or cell.value == "0.00" or cell.value == 0.0:
                    cell.font = Font(color="FFFFFF")  # Solo texto blanco

    wb.save(ruta_archivo)

def comparar_facturas(archivo_odoo, archivo_sat, archivo_salida):
    df_odoo = pd.read_excel(archivo_odoo)
    df_sat = pd.read_excel(archivo_sat)

    # Buscar columnas específicas y manejar errores claros
    def buscar_col(df, palabras, nombre_archivo):
        for palabra in palabras:
            cols = [c for c in df.columns if palabra in c.lower()]
            if cols:
                return cols[0]
        raise Exception(f"No se encontró ninguna columna con alguna de estas palabras: {palabras} en el archivo {nombre_archivo}.\nColumnas detectadas: {list(df.columns)}")

    try:
        col_numero = buscar_col(df_odoo, ['numero'], 'Odoo')
        col_serie = buscar_col(df_sat, ['serie'], 'SAT')
        col_folio = buscar_col(df_sat, ['folio'], 'SAT')
    except Exception as e:
        messagebox.showerror("Error de columnas", str(e))
        return

    # Buscar columnas de estado
    col_estado_pago = None
    col_estado_sat = None
    try:
        col_estado_pago = buscar_col(df_odoo, ['estado pago', 'estado en pago'], 'Odoo')
    except Exception:
        pass
    try:
        col_estado_sat = buscar_col(df_sat, ['estado sat', 'estado'], 'SAT')
    except Exception:
        pass

    # Unir Serie y Folio para formar el número de factura SAT (Folio siempre 5 dígitos con ceros a la izquierda)
    def folio_str(x):
        try:
            x_int = int(float(x))
            return str(x_int).zfill(5)
        except:
            return str(x).strip().zfill(5)
    df_sat['FOLIO_STR'] = df_sat[col_folio].apply(folio_str)
    df_sat['FACTURA'] = df_sat[col_serie].astype(str).str.strip() + df_sat['FOLIO_STR']
    df_sat['FACTURA'] = df_sat['FACTURA'].str.upper()
    df_odoo['FACTURA'] = df_odoo[col_numero].astype(str).str.strip().str.upper()

    set_odoo = set(df_odoo['FACTURA'])
    set_sat = set(df_sat['FACTURA'])

    resultados = []
    for num in sorted(set_odoo):
        estado_pago = ''
        estado_sat = ''
        estado_cancelado = ''
        if col_estado_pago and num in df_odoo['FACTURA'].values:
            estado_pago = str(df_odoo.loc[df_odoo['FACTURA'] == num, col_estado_pago].values[0]).strip().lower()
        if col_estado_sat and num in df_sat['FACTURA'].values:
            estado_sat = str(df_sat.loc[df_sat['FACTURA'] == num, col_estado_sat].values[0]).strip().lower()
        # Estado de cancelado solo si ambos son 'cancelado'
        if estado_pago == 'cancelado' and estado_sat == 'cancelado':
            estado_cancelado = 'TRUE'
        elif estado_pago == 'cancelado' or estado_sat == 'cancelado':
            estado_cancelado = 'FALSE'
        else:
            estado_cancelado = ''
        if num in set_sat:
            resultados.append({'FACTURA ODOO': num, 'FACTURA SAT': num, 'ESTADO DE COMPARACION': 'COINCIDEN', 'ESTADO EN PAGO': estado_pago, 'ESTADO SAT': estado_sat, 'ESTADO CANCELADO': estado_cancelado})
        else:
            resultados.append({'FACTURA ODOO': num, 'FACTURA SAT': '', 'ESTADO DE COMPARACION': 'NO ENCONTRADO EN SAT', 'ESTADO EN PAGO': estado_pago, 'ESTADO SAT': '', 'ESTADO CANCELADO': ''})
    for num in sorted(set_sat - set_odoo):
        estado_sat = ''
        estado_cancelado = ''
        if col_estado_sat and num in df_sat['FACTURA'].values:
            estado_sat = str(df_sat.loc[df_sat['FACTURA'] == num, col_estado_sat].values[0]).strip().lower()
        if estado_sat == 'cancelado':
            estado_cancelado = 'FALSE'
        resultados.append({'FACTURA ODOO': '', 'FACTURA SAT': num, 'ESTADO DE COMPARACION': 'SOLO EN SAT', 'ESTADO EN PAGO': '', 'ESTADO SAT': estado_sat, 'ESTADO CANCELADO': estado_cancelado})

    # Ordenar: primero los cancelados FALSE (rojo pastel), luego TRUE (verde pastel), luego el resto
    df_res = pd.DataFrame(resultados)
    df_res['sort_cancelado'] = df_res['ESTADO CANCELADO'].map({'FALSE': 0, 'TRUE': 1, '': 2})
    df_res['sort_estado'] = df_res['ESTADO DE COMPARACION'].map({'NO ENCONTRADO EN SAT': 0, 'SOLO EN SAT': 1, 'COINCIDEN': 2})
    df_res = df_res.sort_values(by=[ 'sort_estado', 'sort_cancelado', 'FACTURA ODOO', 'FACTURA SAT'], na_position='last').drop(columns=['sort_cancelado','sort_estado'])
    df_res.reset_index(drop=True, inplace=True)

    # Guardar Excel y aplicar colores según reglas:
    # - Coinciden: verde pastel
    # - No coinciden en factura: rojo pastel
    # - Estado cancelado FALSE: rojo pastel
    # - Estado cancelado TRUE: verde pastel
    df_res.to_excel(archivo_salida, index=False)
    wb = openpyxl.load_workbook(archivo_salida)
    ws = wb.active
    rojo_pastel = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
    verde_pastel = PatternFill(start_color="D6FFD6", end_color="D6FFD6", fill_type="solid")
    for row in range(2, ws.max_row+1):
        estado_comparacion = ws.cell(row=row, column=3).value
        estado_cancelado = ws.cell(row=row, column=6).value
        # Estado cancelado FALSE: siempre rojo pastel
        if estado_cancelado == 'FALSE':
            for col in range(1, ws.max_column+1):
                ws.cell(row=row, column=col).fill = rojo_pastel
        # Coinciden o cancelado TRUE: verde pastel
        elif estado_cancelado == 'TRUE' or estado_comparacion == 'COINCIDEN':
            for col in range(1, ws.max_column+1):
                ws.cell(row=row, column=col).fill = verde_pastel
        # No encontrado en SAT: rojo pastel
        elif estado_comparacion == 'NO ENCONTRADO EN SAT':
            for col in range(1, ws.max_column+1):
                ws.cell(row=row, column=col).fill = rojo_pastel
    wb.save(archivo_salida)
    ajustar_formato_excel(archivo_salida)

if __name__ == "__main__":
    app = App()
    app.mainloop()